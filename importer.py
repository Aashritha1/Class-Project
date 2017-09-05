import os
from django.core.wsgi import get_wsgi_application
os.environ['DJANGO_SETTINGS_MODULE'] = 'classproject.settings'
application = get_wsgi_application()
from classproject import settings

from onlineapp import models


from bs4 import BeautifulSoup
import urllib
import openpyxl

import click
import MySQLdb
from MySQLdb import Error

import warnings
warnings.filterwarnings('ignore')
from datetime import date

@click.group()
def cli():
    pass

@cli.command()
def createDB():
    print "Entered create"
    try:
        db = MySQLdb.connect("localhost", settings.DATABASES['default']['USER'], settings.DATABASES['default']['PASSWORD'])
        cursor = db.cursor()
        sql = 'CREATE SCHEMA IF NOT EXISTS ' + settings.DATABASES['default']['NAME']
        cursor.execute(sql)
        db.commit()
    except Error as e:
        print e

@cli.command()
def dropDB():
    print "Entered drop"
    try :
        db = MySQLdb.connect("localhost", settings.DATABASES['default']['USER'], settings.DATABASES['default']['PASSWORD'])
        cursor = db.cursor()
        sql = 'DROP DATABASE ' + settings.DATABASES['default']['NAME']
        cursor.execute(sql)
        db.commit()
    except Error as e:
        print e


def get_data_from_html_page(html_url):
# Loading Html page from the URL:
    html = urllib.urlopen( html_url)
    soup = BeautifulSoup(html, 'html.parser')
    table_tag = soup.table

    th = table_tag.findAll('th')
    th = th[1:]
    tr = table_tag.findAll('tr')
    count = 0
    for tr_tag in tr:
        td = tr_tag.findAll('td')
        td = td[1:]
        temp = []
        student = []
        for td_tag in td:
            temp.append(td_tag.text)
        try:
            name = temp[0]
            keys = name.split('_')
            count += 1
            student.append(models.Student.objects.get(db_folder=keys[2]))
            c = models.MockTest1.objects.create(problem1=temp[1], problem2=temp[2],
                                                problem3=temp[3], problem4=temp[4], total=temp[5],
                                                student=student[0])
            c.save()
        except:
            pass
def load_content_from_excel(excel_path):
    src_wb = openpyxl.load_workbook(excel_path)

    src_ws = src_wb['Colleges'] #load the Colleges Worksheet's DATA

    college_list = []
    students_list = []
    for row in range(2, src_ws.max_row + 1):
        temp = []   #temp[0] = college name , temp[1] = acronym , temp[2] = location , temp[3] = contact
        for col in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=row, column=col)
            temp.append(src_cell.value)
        c = models.College.objects.create(name = temp[0],location = temp[2],
                                          acronym = temp[1],contact = temp[3])
        c.save()

    src_ws = src_wb['Current'] #load the students worksheet's DATA
    for row in range(2, src_ws.max_row + 1):
        temp = []       #temp[0] = name , temp[1] = college , temp[2] = email id , temp[3] = Db_names
        for col in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=row, column=col)
            temp.append(src_cell.value)
        i = models.College.objects.get(acronym = temp[1])
        s = models.Student.objects.create(name = temp[0], dob =date(2017, 06,11), email = temp[2],
                                          db_folder = temp[3].lower(), college = i)
        s.save()

    src_ws = src_wb['Deletions']  # load the students worksheet's DATA
    for row in range(2, src_ws.max_row + 1):
        temp = []  # temp[0] = name , temp[1] = college , temp[2] = email id , temp[3] = Db_names
        for col in range(1, src_ws.max_column + 1):
            src_cell = src_ws.cell(row=row, column=col)
            temp.append(src_cell.value)
        try:
            i = models.College.objects.get(acronym=temp[1])
            s = models.Student.objects.create(name=temp[0], dob=date(2017, 06, 11), email=temp[2],
                                              db_folder=temp[3].lower(), college=i ,dropped_out = 1)
            s.save()
        except:
            pass

@cli.command()
@click.argument('excel_path' , nargs = 1)
@click.argument('html_url', nargs = 1)
def populateData(excel_path , html_url):
    print "Entered populate"
    # database = MySQLdb.connect(host="localhost", user="root", passwd="Lokesh@13", db="classdb")
    # cursor = database.cursor()
    load_content_from_excel(excel_path)
    get_data_from_html_page(html_url)


if __name__ == '__main__':
    # populateData('C:/work/appscourse/djangoproject/students.xlsx' ,"sample.html")
    cli()


"""def get_data_from_html_page():
# Loading Html page from the URL:
    html = urllib.urlopen(
        "https://d1b10bmlvqabco.cloudfront.net/attach/inpg92dp42z2zo/hdff4poirlh7i6/io5hun2sdr21/mock_results.html")
    soup = BeautifulSoup(html, 'html.parser')
    table_tag = soup.table

    th = table_tag.findAll('th')

def load_content_from_excel(excel_path,query, cursor):
    src_wb = openpyxl.load_workbook(excel_path)

    #get the sheet names in a list
    names = src_wb.get_sheet_names()
    for sheet_name in names:
        src_ws = src_wb.get_sheet_by_name(sheet_name)
        for row in range(2, src_ws.max_row + 1):
            temp = []
            for col in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row = row,column = col)
                temp.append(src_cell.value)
            values = (temp[0] , temp[2] , temp[1] , temp[3])
            print values
            cursor.execute(query, values)


def populateData(excel_path, URL):
    pass
    # It creates a sql connection
    try:
        database = MySQLdb.connect(host="localhost", user="root", passwd="Lokesh@13", db="classdb")
    except EOFError:
        pass
    cursor = database.cursor()
    query = 'INSERT INTO onlineapp_college(name, location, acronym, contact) VALUES (%s,%s,%s,%s)'
    load_content_from_excel(excel_path,query, cursor)


    cursor.close()
    database.commit()
    database.close() """